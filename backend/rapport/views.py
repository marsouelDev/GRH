import json
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from drf_spectacular.utils import OpenApiParameter, extend_schema
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from employees.models import Employe
from justification.permissions import IsRhOrAdminRole

from .models import Rapport
from .serializers import RapportSerializer


# ═══════════════════════════════════════════════════════════
# UTILITAIRES
# ═══════════════════════════════════════════════════════════

def _get_employe(user):
    """Récupère l'instance Employe associée à l'utilisateur."""
    if hasattr(user, 'employe'):
        return user.employe
    if isinstance(user, Employe):
        return user
    return Employe.objects.filter(email=user.email).first()


def _parse_donnees(rapport):
    donnees = rapport.donnees or {}
    if isinstance(donnees, str):
        try:
            donnees = json.loads(donnees)
        except json.JSONDecodeError:
            donnees = {}
    return donnees


def _flatten_section(cle, valeur):
    """Aplatit récursivement les données pour un affichage tabulaire propre."""
    rows = []
    label_base = cle.replace('_', ' ').capitalize()

    if isinstance(valeur, dict):
        rows.append((f"▶ {label_base.upper()}", ''))
        for sous_cle, sous_val in valeur.items():
            rows.extend(_flatten_section(sous_cle, sous_val))
    elif isinstance(valeur, list):
        rows.append((f"  {label_base}", ''))
        for item in valeur:
            if isinstance(item, dict):
                parts = ' • '.join(f"{k}: {v}" for k, v in item.items() if v is not None)
                rows.append(('    └', parts))
            else:
                rows.append(('    └', str(item)))
    else:
        rows.append((f"  {label_base}", '' if valeur is None else str(valeur)))

    return rows


def _flatten_donnees(donnees):
    rows = []
    for cle, valeur in donnees.items():
        rows.extend(_flatten_section(cle, valeur))
    return rows


# ═══════════════════════════════════════════════════════════
# VUES PRINCIPALES
# ═══════════════════════════════════════════════════════════

class RapportListCreateView(APIView):
    # ✅ SÉCURITÉ : Seuls les RH et Admin peuvent accéder à cette vue (GET et POST)
    permission_classes = [IsRhOrAdminRole]

    @extend_schema(
        summary="Liste des rapports",
        responses=RapportSerializer(many=True),
        parameters=[OpenApiParameter(name='type', description="Filtrer par type", required=False, type=str)]
    )
    def get(self, request):
        qs = Rapport.objects.select_related('genere_par').order_by('-date_creation')
        type_rapport = request.query_params.get('type')
        if type_rapport:
            qs = qs.filter(type_rapport=type_rapport)
        return Response(RapportSerializer(qs, many=True).data, status=status.HTTP_200_OK)

    @extend_schema(summary="Générer un rapport (RH/Admin uniquement)", request=RapportSerializer, responses=RapportSerializer)
    def post(self, request):
        serializer = RapportSerializer(data=request.data, context={'request': request})
        
        if serializer.is_valid():
            user = request.user
            genere_par = _get_employe(user)
            
            if not genere_par:
                print(f"⚠️ AVERTISSEMENT : Employé non trouvé pour l'utilisateur '{user.username}'. Rapport créé sans auteur.")

            rapport = serializer.save(genere_par=genere_par)
            rapport.genererDonnees()
            
            return Response(RapportSerializer(rapport).data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class RapportDetailView(APIView):
    permission_classes = [IsRhOrAdminRole]

    @extend_schema(summary="Détail d'un rapport", responses=RapportSerializer)
    def get(self, request, pk):
        rapport = get_object_or_404(Rapport.objects.select_related('genere_par'), pk=pk)
        return Response(RapportSerializer(rapport).data, status=status.HTTP_200_OK)

    @extend_schema(summary="Supprimer un rapport", responses=None)
    def delete(self, request, pk):
        rapport = get_object_or_404(Rapport, pk=pk)
        rapport.delete()
        return Response({"detail": "Rapport supprimé définitivement."}, status=status.HTTP_204_NO_CONTENT)


class RapportRegeneView(APIView):
    permission_classes = [IsRhOrAdminRole]

    @extend_schema(summary="Regénérer les données du rapport", responses=RapportSerializer)
    def put(self, request, pk):
        rapport = get_object_or_404(Rapport, pk=pk)
        rapport.genererDonnees()
        return Response(RapportSerializer(rapport).data, status=status.HTTP_200_OK)


# ═══════════════════════════════════════════════════════════
# EXPORT EXCEL (AMÉLIORÉ)
# ═══════════════════════════════════════════════════════════

class ExportRapportExcelView(APIView):
    permission_classes = [IsRhOrAdminRole]

    @extend_schema(summary="Exporter le rapport en Excel")
    def get(self, request, pk):
        rapport = get_object_or_404(Rapport, pk=pk)
        donnees = _parse_donnees(rapport)
        rows = _flatten_donnees(donnees)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport RH"

        # ── 1. En-tête du document ──────────────────────────
        ws.merge_cells('A1:B1')
        ws['A1'] = f"BILAN RH — {rapport.titre.upper()}"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # ── 2. Métadonnées ──────────────────────────────────
        meta = [
            ('Type de rapport', rapport.get_type_rapport_display()),
            ('Période', f"{rapport.date_debut or 'Début'} → {rapport.date_fin or 'Fin'}"),
            ('Généré par', str(rapport.genere_par) if rapport.genere_par else 'Système'),
            ('Date', rapport.date_creation.strftime('%d/%m/%Y à %H:%M')),
        ]
        for i, (label, valeur) in enumerate(meta, start=3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True, color="475569")
            ws.cell(row=i, column=2, value=valeur).font = Font(color="1E293B")

        ws.append([])  # Ligne vide

        # ── 3. En-tête du tableau ───────────────────────────
        header_row = ws.max_row + 1
        ws.cell(row=header_row, column=1, value='INDICATEUR / MÉTRIQUE')
        ws.cell(row=header_row, column=2, value='VALEUR')
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col in (1, 2):
            cell = ws.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[header_row].height = 25

        # ── 4. Données avec style alterné et sections ───────
        for i, (label, valeur) in enumerate(rows):
            row = ws.max_row + 1
            is_section = str(label).strip().startswith('▶')
            is_sub = str(label).strip().startswith('└')
            
            c1 = ws.cell(row=row, column=1, value=label)
            c2 = ws.cell(row=row, column=2, value=valeur)
            c1.border = c2.border = border

            if is_section:
                # Style pour les grandes sections
                c1.font = Font(bold=True, color="FFFFFF", size=10)
                c2.font = Font(bold=True, color="FFFFFF", size=10)
                fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
                c1.fill = c2.fill = fill
                ws.merge_cells(f'A{row}:B{row}')
                c1.alignment = Alignment(horizontal='left', vertical='center')
            else:
                # Style alterné pour la lisibilité
                fill_color = "F1F5F9" if i % 2 == 0 else "FFFFFF"
                c1.fill = c2.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                c1.font = Font(color="334155", size=10)
                c2.font = Font(color="0F172A", size=10, bold=is_sub)
                c1.alignment = Alignment(wrap_text=True)
                c2.alignment = Alignment(wrap_text=True)

        # ── 5. Ajustement des colonnes ──────────────────────
        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 45

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="rapport_{rapport.id}_{rapport.type_rapport}.xlsx"'
        wb.save(response)
        return response


# ═══════════════════════════════════════════════════════════
# EXPORT PDF (AMÉLIORÉ)
# ═══════════════════════════════════════════════════════════

def _pdf_styles():
    base = getSampleStyleSheet()
    return {
        'title': ParagraphStyle(
            'RH_Title', parent=base['Title'],
            fontSize=18, textColor=colors.HexColor('#0F172A'),
            spaceAfter=2, fontName='Helvetica-Bold',
        ),
        'subtitle': ParagraphStyle(
            'RH_Sub', parent=base['Normal'],
            fontSize=11, textColor=colors.HexColor('#4F46E5'),
            spaceAfter=12, fontName='Helvetica-Bold',
        ),
        'meta_label': ParagraphStyle(
            'RH_MetaLabel', parent=base['Normal'],
            fontSize=9, textColor=colors.HexColor('#64748B'),
            fontName='Helvetica-Bold',
        ),
        'meta_value': ParagraphStyle(
            'RH_MetaValue', parent=base['Normal'],
            fontSize=9, textColor=colors.HexColor('#0F172A'),
        ),
        'section': ParagraphStyle(
            'RH_Section', parent=base['Normal'],
            fontSize=10, textColor=colors.white,
            fontName='Helvetica-Bold', spaceAfter=0, spaceBefore=8, leftIndent=5,
        ),
        'cell': ParagraphStyle(
            'RH_Cell', parent=base['Normal'],
            fontSize=9, textColor=colors.HexColor('#334155'),
            wordWrap='CJK', leading=12,
        ),
    }


class ExportRapportPdfView(APIView):
    permission_classes = [IsRhOrAdminRole]

    @extend_schema(summary="Exporter le rapport en PDF")
    def get(self, request, pk):
        rapport = get_object_or_404(Rapport, pk=pk)
        donnees = _parse_donnees(rapport)
        rows = _flatten_donnees(donnees)
        styles = _pdf_styles()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_{rapport.id}_{rapport.type_rapport}.pdf"'

        doc = SimpleDocTemplate(
            response, pagesize=letter,
            rightMargin=15 * mm, leftMargin=15 * mm,
            topMargin=20 * mm, bottomMargin=20 * mm,
        )
        story = []

        # ── 1. Bandeau titre ────────────────────────────────
        story.append(Paragraph(f"BILAN RH", styles['title']))
        story.append(Paragraph(rapport.titre, styles['subtitle']))
        story.append(HRFlowable(width='100%', thickness=2, color=colors.HexColor('#4F46E5'), spaceAfter=15))

        # ── 2. Métadonnées dans un tableau propre ───────────
        meta_data = [
            [Paragraph("Type", styles['meta_label']), Paragraph(rapport.get_type_rapport_display(), styles['meta_value'])],
            [Paragraph("Période", styles['meta_label']), Paragraph(f"{rapport.date_debut or 'Début'} → {rapport.date_fin or 'Fin'}", styles['meta_value'])],
            [Paragraph("Généré par", styles['meta_label']), Paragraph(str(rapport.genere_par) if rapport.genere_par else 'Système', styles['meta_value'])],
            [Paragraph("Date", styles['meta_label']), Paragraph(rapport.date_creation.strftime('%d/%m/%Y à %H:%M'), styles['meta_value'])],
        ]
        meta_table = Table(meta_data, colWidths=[80, 400])
        meta_table.setStyle(TableStyle([
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
            ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#E2E8F0')),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 20))

        # ── 3. Tableau des données ──────────────────────────
        table_rows = []
        for label, valeur in rows:
            is_section = str(label).strip().startswith('▶')
            if is_section:
                clean = str(label).replace('▶', '').strip()
                table_rows.append([
                    Paragraph(clean, styles['section']),
                    Paragraph('', styles['section']),
                ])
            else:
                table_rows.append([
                    Paragraph(str(label), styles['cell']),
                    Paragraph(str(valeur), styles['cell']),
                ])

        if table_rows:
            header = [
                Paragraph('<b>INDICATEUR / MÉTRIQUE</b>', styles['cell']),
                Paragraph('<b>VALEUR</b>', styles['cell']),
            ]
            full_table = Table([header] + table_rows, colWidths=[280, 200], repeatRows=1)

            n = len(table_rows)
            row_styles = []

            for i in range(1, n + 1):
                label_val = rows[i - 1][0] if i - 1 < len(rows) else ''
                is_section = str(label_val).strip().startswith('▶')
                
                if is_section:
                    row_styles.append(('BACKGROUND', (0, i), (1, i), colors.HexColor('#334155')))
                    row_styles.append(('TEXTCOLOR', (0, i), (1, i), colors.white))
                elif i % 2 == 0:
                    row_styles.append(('BACKGROUND', (0, i), (1, i), colors.HexColor('#F8FAFC')))

            full_table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.white),
                ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (1, 0), 10),
                # Global
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                *row_styles,
            ]))
            story.append(full_table)

        # ── 4. Pied de page ─────────────────────────────────
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#CBD5E1')))
        story.append(Spacer(1, 5))
        story.append(Paragraph(
            f"Document généré automatiquement par le système WorkFlow RH — {rapport.date_creation.strftime('%d/%m/%Y')}",
            ParagraphStyle('footer', fontSize=7, textColor=colors.HexColor('#94A3B8'), alignment=1)
        ))

        doc.build(story)
        return response